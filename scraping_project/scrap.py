from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import pandas as pd
import time
import threading
import openai
# import re
# import os.path
from openpyxl import Workbook, load_workbook
class scrap_by_keyword:
    def __init__(self):
        self.base_url = 'https://www.google.com'
        options = Options()
        options.headless = True
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36")
        self.driver = webdriver.Chrome(options=options)
        self.driver.maximize_window()
        self.driver.get(self.base_url)

    def search_keyword(self, keyword):
        try:
            WebDriverWait(self.driver, 60).until(EC.presence_of_element_located((By.XPATH, "//textarea[@id='APjFqb']")))
            user_keyword = self.driver.find_element(By.XPATH, "//textarea[@id='APjFqb']")
            user_keyword.send_keys(keyword)
            user_keyword.submit()
            time.sleep(3)
            
        except:
            pass
    def prompt3(self, prompt2, prompt1, keyword):
        response = ''
        try:
            prompt_save = f"prompt3/{keyword}.xlsx"
            openai.api_key = 'Enter Your API Key'
            try:
                
                wb= load_workbook(prompt_save)
                
            except FileNotFoundError:
                wb = Workbook()

            prompt3 = f""" Prompt 3 conclusion
                            You are a writing assistant and take instructions literally. The output you provide is always the same as the language of the input language. The goal of this prompt is to write a conclusion for the ({keyword}) AND ({prompt1}) and ({prompt2})
                            Write a factual conclusion which will be the end paragraph of a page on a website. Place yourself in the persona of the internet user which is browsing the questions.
                            Use the facts and information which are given to you. Make the conclusion factual and write a summary of the information and questions that are being given to you. Make sure that the output is no longer than 7 sentences and includes the answer on the questions on the page."""
            try:
                response = openai.ChatCompletion.create(
                    model="gpt-4o",  # Specify the GPT-4.0 model
                    messages=[

                        {"role": "user", "content": prompt3},
                    ]
                )
                response = response.choices[0].message['content']
                p2 = wb.active
                p2.cell(row=1, column=1).value = "keyword"
                p2.cell(row=1, column=2).value = "Conclusion"
                next_row = p2.max_row + 1
                p2.cell(row=next_row, column=1).value = keyword
                p2.cell(row=next_row, column=2).value = response
                wb.save(prompt_save)
            
            except:
                pass
            return response
        except:
            return 0
            
    def prompt2(self, prompt1, keyword):
        title_intro = []
        try:
            prompt_save = f"prompt2/{keyword}.xlsx"
            openai.api_key = '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'
            try:
                
                wb= load_workbook(prompt_save)
                
            except FileNotFoundError:
                wb = Workbook()
            
            promt2 = f""" Prompt 2 Title and introduction
                    You are a writing assistant and take instructions literally. The output you provide is always
                    the same as the language of the input language. The goal of this prompt is to write a title and
                    introduction paragraph for ({keyword})
                    The title you're giving is always the ({keyword}).
                    For the introduction text scan and analyze the (All text) for facts statements and relevant
                    information. And also scan and analyze ({prompt1}). Give an introduction
                    which will grab the attention. The answer to the main keyword must be in the first two
                    sentences. And further write the introduction about the page and information that will
                    be answered. Make sure to make the introduction factual and if there is a range in the
                    answers give the range and explain the range based on facts. If it’s logical, put the
                    output of the range in a table.
                    The output must be as below:
                    Title
                    Introduction as output of the giving prompt.

                """
            try:
                response = openai.ChatCompletion.create(
                    model="gpt-4o",  # Specify the GPT-4.0 model
                    messages=[

                        {"role": "user", "content": promt2},
                    ]
                )
                response = response.choices[0].message['content']
                lines = response.strip().split("\n")

                if lines:
                    title = lines[0].strip()
                    introduction = "\n".join(lines[1:]).strip()
                else:
                    title = ""
                    introduction = ""

                title_intro.append(title)
                title_intro.append(introduction)
                p2 = wb.active
                p2.cell(row=1, column=1).value = "Title"
                p2.cell(row=1, column=2).value = "Introduction"
                next_row = p2.max_row + 1
                p2.cell(row=next_row, column=1).value = title
                p2.cell(row=next_row, column=2).value = introduction
                wb.save(prompt_save)
            except:
                pass
        except:
            pass
        return title_intro
    def prompt_question(self, prompts, keyword):
        response = ''
        prompt_ans = []
        try:
            prompt_save = f"prompt/{keyword}.xlsx"
            openai.api_key = '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'
            try:
                
                wbpaw= load_workbook(prompt_save)
                
            except FileNotFoundError:
                wbpaw = Workbook()
            
            prmts = f"""You are a writing assistant and take instructions literally. The output you provide is always the same as the language of the input from the questions and answers. 

                For the answer of question 1 ({prompts[0]}) you are going to analyze ({prompts[1]}). 
                You extract the facts from that text and rewrite it in a Google optimized and plagiarism-free manner. To answer question 1 ({prompts[0]}) in a factual way.

                For the answer of question 2 ({prompts[2]}) you are going to analyze ({prompts[3]}). 
                You extract the facts from that text and rewrite it in a Google optimized and plagiarism-free manner. To answer question 2 ({prompts[2]}) in a factual way.

                For the answer of question 3 ({prompts[4]}) you are going to analyze ({prompts[5]}). 
                You extract the facts from that text and rewrite it in a Google optimized and plagiarism-free manner. To answer question 3({prompts[4]}) in a factual way.

                For the answer of question 4 ({prompts[6]}) you are going to analyze ({prompts[7]}). 
                You extract the facts from that text and rewrite it in a Google optimized and plagiarism-free manner. To answer question 4 ({prompts[6]}) in a factual way.

                For the answer of question 5 ({prompts[8]}) you are going to analyze ({prompts[9]}). 
                You extract the facts from that text and rewrite it in a Google optimized and plagiarism-free manner. To answer question 5({prompts[8]}) in a factual way.

                After the above instruction you are going to give the following output structure. 
                The output always has the structure below, if a question and/or answer is missing, skip it in whole. So in that you skip the question and answer and move on to the next one.  Start the answer with the main answer in the first or second sentence. Never include numbers, * and other signs. 
                {prompts[0]}
                rewrite text as output from Chatgpt answer 1

                {prompts[2]}
                rewrite text as output from Chatgpt answer 2

                {prompts[4]}
                rewrite text as output from Chatgpt answer 3

                {prompts[6]}
                rewrite text as output from Chatgpt answer 4

                {prompts[8]}
                rewrite text as output from Chatgpt answer 5
            """
            try:
                response = openai.ChatCompletion.create(
                    model="gpt-4o",  # Specify the GPT-4.0 model
                    messages=[

                        {"role": "user", "content": prmts},
                    ]
                )
                response = response.choices[0].message['content']
                
                try:
                    segments = response.strip().split("\n\n")
                    questions = []
                    answers = []

                    for segment in segments:

                        index = segment.find('?') + 1
                        question = segment[:index].strip()
                        answer = segment[index:].strip()
                        questions.append(question)
                        answers.append(answer)

                except Exception as e:

                    print(f"index error {e}")

                try:
                    wbpa = wbpaw.active
                    wbpa.cell(row=1, column=1).value = 'Question1'
                    wbpa.cell(row=1, column=2).value = 'Answer1'
                    wbpa.cell(row=1, column=3).value = 'Question2'
                    wbpa.cell(row=1, column=4).value = 'Answer2'
                    wbpa.cell(row=1, column=5).value = 'Question3'
                    wbpa.cell(row=1, column=6).value = 'Answer3'
                    wbpa.cell(row=1, column=7).value = 'Question4'
                    wbpa.cell(row=1, column=8).value = 'Answer4'
                    wbpa.cell(row=1, column=9).value = 'Question5'
                    wbpa.cell(row=1, column=10).value = 'Answer5'
                    next_row = wbpa.max_row + 1
                    for i in range(len(questions)):

                        wbpa.cell(row=next_row, column=i+i+1).value = questions[i]
                        wbpa.cell(row=next_row, column=i+i+2).value = answers[i]
                        prompt_ans.append(questions[i])
                        prompt_ans.append(answers[i])

                    wbpaw.save(prompt_save)

                except Exception as e:

                    print(f"the error in excel is {e}") 

                return prompt_ans
            except Exception as e:
                return 0
        except:
            pass
        
    def scrap_questions(self, keyword):
        questions_arr = []
        question_ans = []
        try:
            WebDriverWait(self.driver, 2).until(EC.presence_of_element_located((By.XPATH,"//div[@jsname='tJHJj']")))
            question_only = f"questions/question_only/{keyword}.xlsx"
            question_answers = f"questions/question_answers/{keyword}.xlsx"
            try:
                
                wbo = load_workbook(question_only)
                wba = load_workbook(question_answers)
            except FileNotFoundError:
                wbo = Workbook()
                wba = Workbook()
            wso = wbo.active
            wsa = wba.active
            
            try:
                self.driver.find_elements(By.XPATH,"//div[@jsname='tJHJj']")[1].click()
                time.sleep(1)
                self.driver.find_elements(By.XPATH,"//div[@jsname='tJHJj']")[1].click()
                questions = self.driver.find_elements(By.XPATH,"//div[@jsname='tJHJj']")
                
                wso.cell(row=1, column=1).value = 'Question1'
                wso.cell(row=1, column=2).value = 'Question2'
                wso.cell(row=1, column=3).value = 'Question3'
                wso.cell(row=1, column=4).value = 'Question4'
                wso.cell(row=1, column=5).value = 'Question5'
                for index, question in enumerate(questions[:6]):
                    question = question.text
                    questions_arr.append(question)
                next_row = wso.max_row + 1
                wso.cell(row=next_row, column=1).value = questions_arr[0]
                wso.cell(row=next_row, column=2).value = questions_arr[1]
                wso.cell(row=next_row, column=3).value = questions_arr[2]
                wso.cell(row=next_row, column=4).value = questions_arr[3]
                wso.cell(row=next_row, column=5).value = questions_arr[4]
                # wso.append(questions_arr)
            except:
                pass
            
            wsa.cell(row=1, column=1).value = 'Question1'
            wsa.cell(row=1, column=2).value = 'Answer1'
            wsa.cell(row=1, column=3).value = 'Question2'
            wsa.cell(row=1, column=4).value = 'Answer2'
            wsa.cell(row=1, column=5).value = 'Question3'
            wsa.cell(row=1, column=6).value = 'Answer3'
            wsa.cell(row=1, column=7).value = 'Question4'
            wsa.cell(row=1, column=8).value = 'Answer4'
            wsa.cell(row=1, column=9).value = 'Question5'
            wsa.cell(row=1, column=10).value = 'Answer5'
            for index, element in enumerate(questions_arr):
                question_ans.append(element)
                try:
                    # self.driver.find_elements(By.XPATH, "//div[@jsname='pcRaIe']")[index].click()
                    # self.driver.find_elements(By.XPATH, "//div[@jsname='tJHJj']")[index].click()
                    self.driver.find_elements(By.XPATH, "//div[@jsname='yEVEwb']")[index].click()
                    
                    qanswer = self.driver.find_elements(By.CLASS_NAME, 'bCOlv')[index].text
                    time.sleep(2)
                    question_ans.append(qanswer)
                except Exception as e:
                    print(f"Failed to click on '{element}': {e}")
            next_row = wsa.max_row + 1
            wsa.cell(row=next_row, column=1).value = question_ans[0]
            wsa.cell(row=next_row, column=2).value = question_ans[1]
            wsa.cell(row=next_row, column=3).value = question_ans[2]
            wsa.cell(row=next_row, column=4).value = question_ans[3]
            wsa.cell(row=next_row, column=5).value = question_ans[4]
            wsa.cell(row=next_row, column=6).value = question_ans[5]
            wsa.cell(row=next_row, column=7).value = question_ans[6]
            wsa.cell(row=next_row, column=8).value = question_ans[7]
            wsa.cell(row=next_row, column=9).value = question_ans[8]
            wsa.cell(row=next_row, column=10).value = question_ans[9]
            wbo.save(question_only)
            wba.save(question_answers)
            return question_ans
        except:
            return 0
    def scrap_all_text(self, link, keyword):
        excel_file_path = f"all_text/{keyword}.xlsx"

        try:

            wb = load_workbook(excel_file_path)

        except FileNotFoundError:

            wb = Workbook()

        ws = wb.active

        all_text = self.driver.find_element(By.TAG_NAME, 'body').text
        ws.cell(row=1, column=1).value = 'link'
        ws.cell(row=1, column=2).value = 'text'

        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1).value = link
        ws.cell(row=next_row, column=2).value = all_text
        wb.save(excel_file_path)


    def save_heading_text(self, keyword, link):
        h1_elements = self.driver.find_elements(By.TAG_NAME, 'h1')
        h2_elements = self.driver.find_elements(By.TAG_NAME, 'h2')
        h3_elements = self.driver.find_elements(By.TAG_NAME, 'h3')


        excel_file_path = f"headings/{keyword}.xlsx"

        try:

            wb = load_workbook(excel_file_path)

        except FileNotFoundError:

            wb = Workbook()

        ws = wb.active
        
        h1_heading=[]
        h2_heading=[]
        h3_heading=[]
        for h1 in h1_elements:
            
            h1_heading.append(h1.text)

        for h2 in h2_elements:

            h2_heading.append(h2.text)

        for h3 in h3_elements:

            h3_heading.append(h3.text)

        ws.cell(row=1, column=1).value = 'link'
        ws.cell(row=1, column=2).value = 'h1'
        ws.cell(row=1, column=3).value = 'h2'
        ws.cell(row=1, column=4).value = 'h3'

        next_row = ws.max_row + 1 
        ws.cell(row=next_row, column=1).value = link

        if h1_heading!='':
            
            ws.cell(row=next_row, column=2).value = ', '.join(map(str, h1_heading))

        if h2_heading!='':

            ws.cell(row=next_row, column=3).value = ', '.join(map(str, h2_heading))

        if h3_heading!='':

            ws.cell(row=next_row, column=4).value = ', '.join(map(str, h3_heading))
        
        wb.save(excel_file_path)

    def scrap_links(self):
        link_elements = self.driver.find_elements(By.CSS_SELECTOR, 'a[jsname="UWckNb"]')
        get_links = []
        
        for index, link_element in enumerate(link_elements[:3]):
            href_value = link_element.get_attribute('href')
            get_links.append(href_value)
           
        return get_links
        
    def visit_unique_link(self, get_links, keyword):

        for index, link in enumerate(get_links):

            self.driver.get(link)
            self.scrap_all_text(link, keyword)
            self.save_heading_text(keyword, link)

        self.driver.get(self.base_url)

    def non_people_asked(self, keyword):

        Nopaq = f"non_People_ask/NPAQ.xlsx"

        try:

            pask = load_workbook(Nopaq)

        except FileNotFoundError:

            pask = Workbook()

        paskq = pask.active
        paskq.cell(row=1, column=1).value = 'Non_asked_question'

        next_row = paskq.max_row + 1
        paskq.cell(row=next_row, column=1).value = keyword
        pask.save(Nopaq)
    
    def final_gpt(self, keyword, final_data):
        final_file = f"final_chatgpt/prompt_final.xlsx"
        try:
            
            pf= load_workbook(final_file)
            
        except FileNotFoundError:
            pf = Workbook()

        p2 = pf.active
        p2.cell(row=1, column=1).value = "Input keyword"
        p2.cell(row=1, column=2).value = "Title"
        p2.cell(row=1, column=3).value = "Introduction"
        p2.cell(row=1, column=4).value = "Question 1"
        p2.cell(row=1, column=5).value = "Answer 1"
        p2.cell(row=1, column=6).value = "Question 2"
        p2.cell(row=1, column=7).value = "Answer 2"
        p2.cell(row=1, column=8).value = "Question 3"
        p2.cell(row=1, column=9).value = "Answer 3"
        p2.cell(row=1, column=10).value = "Question 4"
        p2.cell(row=1, column=11).value = "Answer 4"
        p2.cell(row=1, column=12).value = "Question 5"
        p2.cell(row=1, column=13).value = "Answer 5"
        p2.cell(row=1, column=14).value = "Conclusion"
        next_row = p2.max_row + 1
        p2.cell(row=next_row, column=1).value = keyword
        p2.cell(row=next_row, column=2).value = final_data[1][0]
        p2.cell(row=next_row, column=3).value = final_data[1][1]
        for i in range(len(final_data[0])):
            p2.cell(row=next_row, column=i+4).value = final_data[0][i]
        p2.cell(row=next_row, column=14).value = final_data[2]
        pf.save(final_file)


    def close_driver(self):

        try:
            self.driver.quit()
        except Exception as e:
            print(f"Failed to close driver: {e}")
        
def scrape_in_thread(keyword, bot):

    try:
        final_csv = []
        bot.search_keyword(keyword)
        question_answers = bot.scrap_questions(keyword)
        try:
            if question_answers!=0:
                promptqa = bot.prompt_question(question_answers, keyword)
                final_csv.append(promptqa)
                # promptqa = ", ".join(promptqa)
                prompt_title_intro = bot.prompt2(promptqa, keyword)
                final_csv.append(prompt_title_intro)
                # prompt_title_intro = ", ".join(prompt_title_intro)
                conclusion = bot.prompt3(prompt_title_intro, promptqa, keyword)
                final_csv.append(conclusion)
                bot.final_gpt(keyword, final_csv)
            else:
                bot.non_people_asked(keyword)
        except Exception as e:
            print(f"Exception in prompts {e}")

        links = bot.scrap_links()
        bot.visit_unique_link(links, keyword)
        
    except Exception as e:

        print(f"Error in thread for '{keyword}': {e}")

    finally:

        bot.close_driver()

def run_threads(df, batch_size):
    try:
        i = 0
        threads = []
        bots = []

        while i < len(df):
            
            for _ in range(batch_size):
                if i < len(df):
                    row = df.iloc[i]
                    bot = scrap_by_keyword()  # Assuming scrap_by_keyword() returns a bot instance
                    bots.append(bot)
                    keyword = row['keyword']
                    thread = threading.Thread(target=scrape_in_thread, args=(keyword, bot))
                    threads.append(thread)
                    thread.start()
                    i += 1
            
            
            for thread in threads:
                thread.join()
            
           
            for bot in bots:
                bot.close_driver()
            threads.clear()
            bots.clear()

        print("All batches completed.")

    except Exception as e:
        print(f"Main program error: {e}")

if __name__ == "__main__":
    try:
        df = pd.read_csv('prompt.csv')  # Assuming 'prompt.csv' contains 'keyword' column
        batch_size = int(input("Enter Number of Threads:"))
        run_threads(df, batch_size)

    except Exception as e:
        print(f"Main program error: {e}")